import sys
import time

import openpyxl
from openpyxl import Workbook
from docx2python import docx2python
import re
from openpyxl.styles import Font, PatternFill


class ReadDocFile:
    def __init__(self, word_doc_name):
        self.__word_doc_name = word_doc_name
        self.__new_txt_name = "docx_to_txt.txt"

    def doc_to_txt(self):
        print("Converting the file "+self.__word_doc_name+" to text file "+ self.__new_txt_name)
        time.sleep(2)
        document = docx2python(docx_filename=self.__word_doc_name)
        file = open(self.__new_txt_name, "w", errors="ignore")

        file.writelines(document.text)

        return self.__new_txt_name


class ReadTxtFileForFootnotes:
    def __init__(self, doc_txt_file, reference_section_count_up_to_wich_word="Eidesstattliche"):
        self.__doc_txt_file = doc_txt_file
        self.__xlxs_footnote_workbook = "footnotes_file.xlsx"
        self.__reference_section_count_up_to_wich_word = reference_section_count_up_to_wich_word

        self.footnote_workbook = Workbook()
        self.footnote_worksheet = self.footnote_workbook.active
        self.__line_couner = 1

        self.footnote_worksheet.column_dimensions['A'].width = 11

        self.footnote_worksheet.column_dimensions['B'].width = 124
        self.footnote_worksheet.column_dimensions['C'].width = 11
        self.footnote_worksheet.column_dimensions['D'].width = 35
        self.footnote_worksheet.column_dimensions['E'].width = 130

        # write content to the header columns
        self.footnote_worksheet['A1'] = "FN Number"
        self.footnote_worksheet['B1'] = "Footnote text"
        self.footnote_worksheet['C1'] = "Ref Number"
        self.footnote_worksheet['D1'] = "Reference section 1. Word"
        self.footnote_worksheet['E1'] = "Text from reference"

        # set alignment for cell A1
        self.footnote_worksheet['A1'].alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )

        # Set header items to big font
        self.footnote_worksheet['A1'].font = Font(bold=True)
        self.footnote_worksheet['B1'].font = Font(bold=True)
        self.footnote_worksheet['C1'].font = Font(bold=True)
        self.footnote_worksheet['D1'].font = Font(bold=True)
        self.footnote_worksheet['E1'].font = Font(bold=True)

        # Set background color for the header items
        fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        self.footnote_worksheet['A1'].fill = fill
        self.footnote_worksheet['B1'].fill = fill
        self.footnote_worksheet['C1'].fill = fill
        self.footnote_worksheet['D1'].fill = fill
        self.footnote_worksheet['E1'].fill = fill

    def read_foot_notes(self):

        print("Read Footnotes from docx file and store it footnotes_file.xlsx ...")
        time.sleep(2)
        with open(self.__doc_txt_file) as file:

            for line in file:

                footnote = re.search("^(footnote)", line)
                if (footnote != None):
                    self.__line_couner = self.__line_couner + 1
                    foot_note_id = line.split(")")[0]
                    footnote_number = foot_note_id.split("footnote")[1]
                    footnote_text = line.split(")")[1]

                    self.footnote_worksheet['A' + str(self.__line_couner)] = int(footnote_number)
                    self.footnote_worksheet['A' + str(self.__line_couner)].alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center"
                    )

                    self.footnote_worksheet['B' + str(self.__line_couner)] = footnote_text.replace("Vgl. ", "")
                    print("[B"+str(self.__line_couner)+"] = " + footnote_text.replace("Vgl. ", ""))


        return self.__line_couner

    def read_references(self):

        print("Looking for references in the reference section programm will stop before finding the word :"+ self.__reference_section_count_up_to_wich_word)
        time.sleep(2)

        referece_found = False
        counter = 1
        insert_references_in_line = 2

        with open(self.__doc_txt_file) as file:

            for line in file:

                reference_section = re.search("^(Literaturverzeichnis)", line)

                if (reference_section != None and referece_found == False):
                    referece_found = True
                    continue

                if (referece_found):

                    if line.split(" ")[0] == "<a":
                        continue

                    if (line.split(" ")[0]):

                        new_line = line.split(" ")

                        if (new_line):
                            counter = counter + 1

                            if new_line[0] != '\n':

                                if new_line[0] == self.__reference_section_count_up_to_wich_word:
                                    break
                                else:

                                    print("[C"+str(insert_references_in_line)+"] = " + ",".join(new_line))
                                    self.footnote_worksheet['C' + str(insert_references_in_line)] = int(insert_references_in_line-1)

                                    self.footnote_worksheet['C' + str(insert_references_in_line)].alignment = openpyxl.styles.Alignment(
                                        horizontal="center", vertical="center"
                                    )

                                    self.footnote_worksheet['D' + str(insert_references_in_line)] = new_line[0]
                                    self.footnote_worksheet['E' + str(insert_references_in_line)] = ",".join(new_line)

                                    insert_references_in_line = insert_references_in_line + 1


        self.footnote_workbook.save(self.__xlxs_footnote_workbook)

        return insert_references_in_line


if __name__ == '__main__':

    word_document = "Implementierung_reading_footnotes.docx"

    convert_doc_to_txt = ReadDocFile(word_document)

    try:
        txt_file_name = convert_doc_to_txt.doc_to_txt()
    except:
        sys.exit("ERROR: docx file is already open please close it ")

    read_txt_file_for_footnotes = ReadTxtFileForFootnotes(
        doc_txt_file=txt_file_name,
        reference_section_count_up_to_wich_word="Eidesstattliche"
    )

    line_counter_footnotes = read_txt_file_for_footnotes.read_foot_notes()

    try:
        line_counter_references = read_txt_file_for_footnotes.read_references()
    except:
        sys.exit("ERROR: xlsx file is already open please close it ")

    if line_counter_footnotes != line_counter_references:
        print("FAILED: Number of footnotes and number of References not matching")
        print("Number if footnotes  = ",line_counter_footnotes-1)
        print("Number if references = ",line_counter_references-1)

    elif line_counter_footnotes == line_counter_references:
        print("CORRECT: Number of footnotes matching the number of references in the reverence section")